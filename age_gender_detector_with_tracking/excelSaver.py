import openpyxl
from datetime import datetime

class excelSaver:
    def __init__(self, fileName):
        self.wb = openpyxl.Workbook()
        self.fileName = fileName
        
        now = datetime.now()
        
        self.ws = self.wb.create_sheet("%s-%s-%s" %(now.year, now.month, now.day))
        self.ws.cell(1, column = 1).value = "objectID"
        self.ws.cell(1, column = 2).value = "date"
        self.ws.cell(1, column = 3).value = "age"
        self.ws.cell(1, column = 4).value = "gender"
        self.wb.save("./output/%s"%self.fileName)
        self.indexer = 2
        
    def regist(self, objectID, age, gender):
        self.ws.cell(self.indexer, column = 1).value = objectID
        self.ws.cell(self.indexer, column = 2).value = datetime.now()
        self.ws.cell(self.indexer, column = 3).value = age
        self.ws.cell(self.indexer, column = 4).value = gender
        self.wb.save("./output/%s"%self.fileName)
        self.indexer += 1
